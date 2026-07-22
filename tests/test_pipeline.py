from pathlib import Path

import pytest
from openpyxl import Workbook
from pypdf import PdfReader

from historical_bloodlines.application.services.build_genealogy import (
    BuildGenealogyUseCase,
    PageFormat,
)


HEADERS = [
    "№",
    "Имя",
    "Титул",
    "Начало правления",
    "Конец правления",
    "Дети",
    "Брак",
]


def test_excel_to_multi_page_pdf(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.pdf"

    workbook = Workbook()
    first = workbook.active
    first.title = "First dynasty"
    first.append(HEADERS)
    first.append([1, "Parent A", "King", 1000, 1020, "Child", "Parent B"])
    first.append([2, "Parent B", "Queen", 1001, 1030, "Child", "Parent A"])
    first.append([3, "Child", "Prince", None, None, None, None])

    second = workbook.create_sheet("Second dynasty")
    second.append(HEADERS)
    second.append([1, "Ancestor", None, None, None, "Descendant", None])
    second.append([2, "Descendant", None, None, None, None, None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert rendered == output
    assert rendered.exists()
    pages = PdfReader(str(rendered)).pages
    assert len(pages) == 2
    for page in pages:
        assert abs(float(page.mediabox.width) - 595.2756) < 0.1
        assert abs(float(page.mediabox.height) - 419.5276) < 0.1
    assert warnings == ()


def test_pdf_can_be_rendered_as_a4(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree_a4.pdf"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Parent", None, None, None, "Child", None])
    sheet.append([2, "Child", None, None, None, None, None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(
        source,
        output,
        page_format=PageFormat.A4,
    )

    page = PdfReader(str(rendered)).pages[0]
    assert abs(float(page.mediabox.width) - 841.8898) < 0.1
    assert abs(float(page.mediabox.height) - 595.2756) < 0.1
    assert warnings == ()


def test_excel_to_separate_svg_files(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.svg"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Parent", None, None, None, "Child", None])
    sheet.append([2, "Child", None, None, None, None, None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert rendered.is_dir()
    assert len(tuple(rendered.glob("*.svg"))) == 1
    assert warnings == ()


def test_ambiguous_names_are_resolved_by_qualifier_and_row_context(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.pdf"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Ancestor", None, None, None, "Albert II (duke)", None])
    sheet.append([2, "Albert II", "duke", None, None, None, None])
    sheet.append([3, "Other", None, None, None, "Albert II", None])
    sheet.append([4, "Albert II", "king", None, None, None, None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert rendered.exists()
    assert warnings == ()


def test_terminal_branch_label_becomes_placeholder_node(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.pdf"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Ancestor", None, None, None, "Counts of Example", None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert rendered.exists()
    assert len(warnings) == 1
    assert "created placeholder" in warnings[0]
    assert "Counts of Example" in warnings[0]


def test_render_is_landscape_without_visible_marriage_points(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    pdf_output = tmp_path / "tree.pdf"
    svg_output = tmp_path / "tree.svg"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Ancestor", None, None, None, "Matilda", None])
    sheet.append([2, "First spouse", None, None, None, None, "Matilda"])
    sheet.append([3, "Matilda", None, None, None, "Child", "First spouse; Second spouse"])
    sheet.append([4, "Second spouse", None, None, None, "Child", "Matilda"])
    sheet.append([5, "Child", None, None, None, None, None])
    workbook.save(source)

    pdf_path, warnings = BuildGenealogyUseCase().execute(source, pdf_output)
    assert warnings == ()
    page = PdfReader(str(pdf_path)).pages[0]
    assert float(page.mediabox.width) > float(page.mediabox.height)

    svg_directory, warnings = BuildGenealogyUseCase().execute(source, svg_output)
    assert warnings == ()
    svg_text = next(svg_directory.glob("*.svg")).read_text(encoding="utf-8")

    # A person participating in two marriages is rendered once and reused.
    assert svg_text.count(">Matilda<") == 1
    # Marriage/family junctions exist only as invisible anchors.
    assert "<ellipse" not in svg_text
    assert "<circle" not in svg_text
    # All visible relationship lines use one stroke width.
    assert 'stroke-width="0.' not in svg_text
    assert 'font-family="Sans"' in svg_text


def test_children_are_not_automatically_assigned_to_unconfirmed_spouse(tmp_path: Path) -> None:
    from historical_bloodlines.application.services.assembler import GenealogyAssembler
    from historical_bloodlines.application.services.parser import GenealogyRowParser
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader

    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Parent", None, None, None, "Child", "Second spouse"])
    sheet.append([2, "Second spouse", None, None, None, None, "Parent"])
    sheet.append([3, "Child", None, None, None, None, None])
    workbook.save(source)

    sheet_dto = ExcelGenealogyReader().read(source)[0]
    parser = GenealogyRowParser()
    genealogy = GenealogyAssembler().assemble(parser.parse(row) for row in sheet_dto.rows)

    child = next(person for person in genealogy.persons.values() if person.name == "Child")
    child_families = [
        relation
        for relation in genealogy.family_child_relations
        if relation.child_id == child.id
    ]
    assert len(child_families) == 1
    assert len(child_families[0].parent_ids) == 1


def test_child_reference_does_not_resolve_to_same_named_spouse(tmp_path: Path) -> None:
    from historical_bloodlines.application.services.assembler import GenealogyAssembler
    from historical_bloodlines.application.services.parser import GenealogyRowParser
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader

    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Adele", None, None, None, "Stephen", "Stephen"])
    sheet.append([2, "Stephen", "count", None, None, "Stephen", "Adele"])
    sheet.append([3, "Stephen", "king", 1135, 1154, None, None])
    workbook.save(source)

    sheet_dto = ExcelGenealogyReader().read(source)[0]
    parser = GenealogyRowParser()
    genealogy = GenealogyAssembler().assemble(parser.parse(row) for row in sheet_dto.rows)

    people = sorted(
        (person for person in genealogy.persons.values() if person.name == "Stephen"),
        key=lambda person: person.source_key.row_number,
    )
    spouse, child = people
    adele = next(person for person in genealogy.persons.values() if person.name == "Adele")

    assert any(
        relation.parent_id == adele.id and relation.child_id == child.id
        for relation in genealogy.parent_child_relations
    )
    assert not any(
        relation.parent_id == adele.id and relation.child_id == spouse.id
        for relation in genealogy.parent_child_relations
    )


def test_svg_relationship_segments_share_exact_junctions(tmp_path: Path) -> None:
    from historical_bloodlines.infrastructure.graph.renderer import GraphvizGenealogyRenderer

    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.svg"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Parent A", None, None, None, "Child", "Parent B"])
    sheet.append([2, "Parent B", None, None, None, "Child", "Parent A"])
    sheet.append([3, "Child", None, None, None, None, None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert warnings == ()
    svg_text = next(rendered.glob("*.svg")).read_text(encoding="utf-8")
    assert 'stroke="#222222"' in svg_text
    assert 'marker-end=' not in svg_text

    # A horizontal bus is cut exactly at every T-junction. Adjacent pieces
    # share the same endpoint and never extend beyond the requested geometry.
    segments = GraphvizGenealogyRenderer._horizontal_bus_segments(
        20.0,
        [30.0, 10.0, 20.0, 20.0],
    )
    assert segments == (
        (10.0, 20.0, 20.0, 20.0),
        (20.0, 20.0, 30.0, 20.0),
    )


def test_single_child_component_is_snapped_under_family_source(tmp_path: Path) -> None:
    from historical_bloodlines.application.services.assembler import GenealogyAssembler
    from historical_bloodlines.application.services.parser import GenealogyRowParser
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader
    from historical_bloodlines.infrastructure.graph.renderer import GraphvizGenealogyRenderer

    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append(HEADERS)
    sheet.append([1, "Parent A", None, None, None, "Only child", "Parent B"])
    sheet.append([2, "Parent B", None, None, None, "Only child", "Parent A"])
    sheet.append([3, "Only child", None, None, None, "Grandchild", "Spouse"])
    sheet.append([4, "Spouse", None, None, None, None, "Only child"])
    sheet.append([5, "Grandchild", None, None, None, None, None])
    workbook.save(source)

    sheet_dto = ExcelGenealogyReader().read(source)[0]
    parser = GenealogyRowParser()
    genealogy = GenealogyAssembler().assemble(parser.parse(row) for row in sheet_dto.rows)

    renderer = GraphvizGenealogyRenderer()
    components, component_by_person = renderer._build_partner_components(genealogy)
    families = renderer._build_families(genealogy, component_by_person, components)
    component_graph = renderer._build_component_graph(components, families, component_by_person)

    levels: dict[int, int] = {}
    import networkx as nx
    for component_id in nx.topological_sort(component_graph):
        predecessors = tuple(component_graph.predecessors(component_id))
        levels[component_id] = (
            max(
                (levels[parent_id] for parent_id in predecessors),
                default=-1,
            )
            + 1
        )
    for component_id in components:
        levels.setdefault(component_id, 0)

    centers = renderer._place_tree_components(
        genealogy,
        components,
        component_graph,
        component_by_person,
        families,
    )

    target_family = next(
        family
        for family in families
        if len(family.child_ids) == 1 and len(family.parent_ids) == 2
    )
    child_id = target_family.child_ids[0]
    child_component_id = component_by_person[child_id]

    disturbed_centers = dict(centers)
    disturbed_centers[child_component_id] += 10.0
    adjusted = renderer._realign_single_child_components(
        genealogy,
        components,
        component_graph,
        component_by_person,
        families,
        levels,
        disturbed_centers,
    )

    desired_center = (
        adjusted[target_family.parent_component_id]
        + target_family.source_offset
        - components[child_component_id].person_offsets[child_id]
    )
    assert abs(adjusted[child_component_id] - desired_center) < 0.1


def test_sheet_title_is_read_from_special_nazvanie_row(tmp_path: Path) -> None:
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader

    source = tmp_path / "input.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Short sheet name"
    sheet.append(["Название", "Полное название родословной"])
    sheet.append(HEADERS)
    sheet.append([1, "Ancestor", None, None, None, None, None])
    workbook.save(source)

    parsed_sheet = ExcelGenealogyReader().read(source)[0]

    assert parsed_sheet.name == "Short sheet name"
    assert parsed_sheet.display_title == "Полное название родословной"
    assert parsed_sheet.rows[0].source_sheet == "Short sheet name"


def test_render_uses_special_nazvanie_row_for_pdf_title_and_filename(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.svg"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet internal name"
    sheet.append(["Название", "Династия Комнинов"])
    sheet.append(HEADERS)
    sheet.append([1, "Parent", None, None, None, "Child", None])
    sheet.append([2, "Child", None, None, None, None, None])
    workbook.save(source)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert warnings == ()
    output_files = tuple(rendered.glob("*.svg"))
    assert len(output_files) == 1
    assert output_files[0].name.startswith("001_Династия_Комнинов")
    svg_text = output_files[0].read_text(encoding="utf-8")
    assert "Династия Комнинов" in svg_text


def test_graphviz_is_invoked_with_ascii_temporary_filename(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from graphviz import Graph

    source = tmp_path / "input.xlsx"
    output = tmp_path / "tree.svg"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet internal name"
    sheet.append(["Название", "Династия Комнинов"])
    sheet.append(HEADERS)
    sheet.append([1, "Parent", None, None, None, "Child", None])
    sheet.append([2, "Child", None, None, None, None, None])
    workbook.save(source)

    invoked_filenames: list[str] = []

    def fake_render(self, *, filename, directory, cleanup, neato_no_op):
        invoked_filenames.append(filename)
        rendered = Path(directory) / f"{filename}.{self.format}"
        rendered.write_text("<svg></svg>", encoding="utf-8")
        return str(rendered)

    monkeypatch.setattr(Graph, "render", fake_render)

    rendered, warnings = BuildGenealogyUseCase().execute(source, output)

    assert warnings == ()
    assert len(invoked_filenames) == 1
    assert invoked_filenames[0].isascii()
    final_output = rendered / "001_Династия_Комнинов.svg"
    assert final_output.is_file()


def test_optional_generation_and_order_columns_control_layout(tmp_path: Path) -> None:
    from historical_bloodlines.application.services.assembler import GenealogyAssembler
    from historical_bloodlines.application.services.parser import GenealogyRowParser
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader
    from historical_bloodlines.infrastructure.graph.renderer import GraphvizGenealogyRenderer

    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append([*HEADERS, "Поколение", "Порядок в поколении"])
    sheet.append([1, "Root", None, None, None, "Right; Left", None, 1, 10])
    # Workbook row order is intentionally opposite to the requested visual order.
    sheet.append([2, "Right", None, None, None, None, None, 3, 20])
    sheet.append([3, "Left", None, None, None, None, None, 3, 10])
    workbook.save(source)

    sheet_dto = ExcelGenealogyReader().read(source)[0]
    parser = GenealogyRowParser()
    genealogy = GenealogyAssembler().assemble(parser.parse(row) for row in sheet_dto.rows)

    renderer = GraphvizGenealogyRenderer()
    components, component_by_person = renderer._build_partner_components(genealogy)
    families = renderer._build_families(genealogy, component_by_person, components)
    component_graph = renderer._build_component_graph(
        components,
        families,
        component_by_person,
    )
    centers, levels = renderer._place_components(
        genealogy,
        components,
        component_graph,
        component_by_person,
        families,
    )

    people = {person.name: person for person in genealogy.persons.values()}
    left_component = component_by_person[people["Left"].id]
    right_component = component_by_person[people["Right"].id]

    assert levels[left_component] == 2
    assert levels[right_component] == 2
    assert centers[left_component] < centers[right_component]


def test_conflicting_generations_inside_partnership_are_rejected(tmp_path: Path) -> None:
    from historical_bloodlines.application.services.assembler import GenealogyAssembler
    from historical_bloodlines.application.services.parser import GenealogyRowParser
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader
    from historical_bloodlines.infrastructure.graph.renderer import GraphvizGenealogyRenderer

    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append([*HEADERS, "Поколение", "Порядок в поколении"])
    sheet.append([1, "First", None, None, None, None, "Second", 1, 10])
    sheet.append([2, "Second", None, None, None, None, "First", 2, 20])
    workbook.save(source)

    sheet_dto = ExcelGenealogyReader().read(source)[0]
    parser = GenealogyRowParser()
    genealogy = GenealogyAssembler().assemble(parser.parse(row) for row in sheet_dto.rows)

    with pytest.raises(ValueError, match="same generation"):
        GraphvizGenealogyRenderer()._build_partner_components(genealogy)


def test_child_cannot_be_forced_above_minimum_generation(tmp_path: Path) -> None:
    from historical_bloodlines.application.services.assembler import GenealogyAssembler
    from historical_bloodlines.application.services.parser import GenealogyRowParser
    from historical_bloodlines.infrastructure.excel import ExcelGenealogyReader
    from historical_bloodlines.infrastructure.graph.renderer import GraphvizGenealogyRenderer

    source = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dynasty"
    sheet.append([*HEADERS, "Поколение", "Порядок в поколении"])
    sheet.append([1, "Parent", None, None, None, "Child", None, 2, 10])
    sheet.append([2, "Child", None, None, None, None, None, 2, 10])
    workbook.save(source)

    sheet_dto = ExcelGenealogyReader().read(source)[0]
    parser = GenealogyRowParser()
    genealogy = GenealogyAssembler().assemble(parser.parse(row) for row in sheet_dto.rows)
    renderer = GraphvizGenealogyRenderer()
    components, component_by_person = renderer._build_partner_components(genealogy)
    families = renderer._build_families(genealogy, component_by_person, components)
    component_graph = renderer._build_component_graph(
        components,
        families,
        component_by_person,
    )

    with pytest.raises(ValueError, match="too early"):
        renderer._place_components(
            genealogy,
            components,
            component_graph,
            component_by_person,
            families,
        )
