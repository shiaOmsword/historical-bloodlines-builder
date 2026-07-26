from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from historical_bloodlines.application.dto import GenealogySheetDTO, RawGenealogyRowDTO


class ExcelGenealogyReader:
    REQUIRED_HEADERS = {
        "№": "row_number",
        "Имя": "person_name",
        "Титул": "title_raw",
        "Начало правления": "reign_start_raw",
        "Конец правления": "reign_end_raw",
        "Дети": "children_raw",
        "Брак": "spouses_raw",
    }
    OPTIONAL_HEADERS = {
        "Поколение": "generation_raw",
        "Порядок в поколении": "generation_order_raw",
    }
    TITLE_HEADER = "Название"

    def read(self, path: Path) -> tuple[GenealogySheetDTO, ...]:
        if not path.exists():
            raise FileNotFoundError(path)
        if path.suffix.casefold() != ".xlsx":
            raise ValueError("Only .xlsx files are supported")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets: list[GenealogySheetDTO] = []
        try:
            for worksheet in workbook.worksheets:
                display_title, rows = self._read_sheet(worksheet)
                if rows:
                    sheets.append(
                        GenealogySheetDTO(
                            name=worksheet.title,
                            display_title=display_title,
                            rows=rows,
                        )
                    )
        finally:
            workbook.close()
        return tuple(sheets)

    def _read_sheet(self, worksheet) -> tuple[str, tuple[RawGenealogyRowDTO, ...]]:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            first_row = next(iterator)
        except StopIteration:
            return worksheet.title, ()

        display_title = worksheet.title
        headers = first_row
        if self._is_title_row(first_row):
            display_title = self._extract_display_title(first_row, worksheet.title)
            try:
                headers = next(iterator)
            except StopIteration:
                return display_title, ()

        header_index = {
            str(value).strip(): index
            for index, value in enumerate(headers)
            if value is not None
        }
        missing = set(self.REQUIRED_HEADERS) - set(header_index)
        if missing:
            raise ValueError(
                f"Sheet {worksheet.title!r} misses headers: {sorted(missing)}"
            )

        rows: list[RawGenealogyRowDTO] = []
        seen_source_numbers: dict[int, tuple[int, str]] = {}
        data_start_row = 3 if self._is_title_row(first_row) else 2
        for excel_row_number, values in enumerate(iterator, start=data_start_row):
            person_name = self._value(values, header_index["Имя"])
            if person_name is None or not str(person_name).strip():
                continue
            person_name_text = str(person_name).strip()

            source_number = self._value(values, header_index["№"])
            row_number = self._parse_source_number(
                source_number,
                worksheet_title=worksheet.title,
                excel_row_number=excel_row_number,
                person_name=person_name_text,
            )
            if row_number is None:
                row_number = excel_row_number

            previous = seen_source_numbers.get(row_number)
            if previous is not None:
                previous_excel_row, previous_name = previous
                raise ValueError(
                    f"Лист {worksheet.title!r}: повторяющийся № {row_number} "
                    f"в ячейках A{previous_excel_row} ({previous_name!r}) и "
                    f"A{excel_row_number} ({person_name_text!r}). "
                    "Значения в колонке '№' должны быть уникальными."
                )
            seen_source_numbers[row_number] = (excel_row_number, person_name_text)

            rows.append(
                RawGenealogyRowDTO(
                    row_number=row_number,
                    source_sheet=worksheet.title,
                    person_name=person_name_text,
                    title_raw=self._optional_text(values, header_index["Титул"]),
                    reign_start_raw=self._value(
                        values,
                        header_index["Начало правления"],
                    ),
                    reign_end_raw=self._value(
                        values,
                        header_index["Конец правления"],
                    ),
                    children_raw=self._optional_text(values, header_index["Дети"]),
                    spouses_raw=self._optional_text(values, header_index["Брак"]),
                    generation_raw=self._optional_value(
                        values,
                        header_index.get("Поколение"),
                    ),
                    generation_order_raw=self._optional_value(
                        values,
                        header_index.get("Порядок в поколении"),
                    ),
                )
            )
        return display_title, tuple(rows)


    @staticmethod
    def _parse_source_number(
        value: object | None,
        *,
        worksheet_title: str,
        excel_row_number: int,
        person_name: str,
    ) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            raise ValueError(
                f"Лист {worksheet_title!r}, ячейка A{excel_row_number} "
                f"({person_name!r}): значение '№' должно быть целым числом."
            )
        try:
            number = int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Лист {worksheet_title!r}, ячейка A{excel_row_number} "
                f"({person_name!r}): значение '№' должно быть целым числом."
            ) from exc
        if isinstance(value, float) and not value.is_integer():
            raise ValueError(
                f"Лист {worksheet_title!r}, ячейка A{excel_row_number} "
                f"({person_name!r}): значение '№' должно быть целым числом."
            )
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped or str(number) != stripped:
                raise ValueError(
                    f"Лист {worksheet_title!r}, ячейка A{excel_row_number} "
                    f"({person_name!r}): значение '№' должно быть целым числом."
                )
        if number < 1:
            raise ValueError(
                f"Лист {worksheet_title!r}, ячейка A{excel_row_number} "
                f"({person_name!r}): значение '№' должно быть положительным."
            )
        return number

    @classmethod
    def _is_title_row(cls, values: tuple[object, ...]) -> bool:
        if not values:
            return False
        first = values[0]
        return first is not None and str(first).strip() == cls.TITLE_HEADER

    @classmethod
    def _extract_display_title(cls, values: tuple[object, ...], fallback: str) -> str:
        for value in values[1:]:
            if value is not None and str(value).strip():
                return str(value).strip()
        return fallback

    @staticmethod
    def _value(values: tuple[object, ...], index: int) -> object | None:
        return values[index] if index < len(values) else None

    @classmethod
    def _optional_text(cls, values: tuple[object, ...], index: int) -> str | None:
        value = cls._value(values, index)
        return None if value in (None, "") else str(value)

    @classmethod
    def _optional_value(
        cls,
        values: tuple[object, ...],
        index: int | None,
    ) -> object | None:
        if index is None:
            return None
        return cls._value(values, index)
